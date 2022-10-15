import PySimpleGUI as sg
from pathlib import Path
from openpyxl import load_workbook
from utils import get_kids, get_work_days, date
from conf import *
from loguru import logger


def mark_kids(file_name: str):
    """
    Создает окно для проверки отсутствующих
    :param file_name: str имя файла без расширения
    """
    month = file_name.split('_')[-1][:-5]
    work_days = get_work_days(month)

    path = Path(Path.cwd(), 'Ведомости', file_name)

    # Получаем список детей из файла Excel
    kids = get_kids(path)
    # Получаем отсутствующих
    absent = get_absent(path, f'{date:%d}', kids)
    # Левая колонка: текстовые виджеты с именами детей из списка kids
    left_col = [[sg.Text(kid)] for kid in kids]
    # Правая колонка с пустыми инпутами, пронумерованные от нуля
    right_col = [[sg.Input(size=(2, 1), key=i, default_text=absent[i])] for i in range(len(kids))]
    # Устанавливаем фокус на первого ребенка
    right_col[0][0].Focus = True

    layout = [
        [sg.Text(f'{file_name}')],
        [sg.Frame('Дата',
                  [[sg.Text('Число: '),
                    sg.Combo(work_days, default_value=f'{date:%d}', k='date', enable_events=True),
                    sg.Text(month)]])],
        [sg.Column(left_col), sg.Column(right_col)],
        [sg.Button('Отметить')]
    ]

    window = sg.Window('Отметить ученика', layout, finalize=True, element_justification='center',
                       return_keyboard_events=True, font=(FONT_FAMILY, FONT_SIZE))
    # При фокусе на кнопке с ключом "Отметить", нажатие Enter генерирует событие "Отметить_Enter"
    window['Отметить'].bind("<Return>", "_Enter")

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        focus = window.find_element_with_focus()
        prev_focus = focus.get_previous_focus()
        next_focus = focus.get_next_focus()
        if event == 'date':
            absent = get_absent(path, values['date'], kids)
            for i in range(len(absent)):
                window[i].Update(absent[i])
        if event == 'Right:39' and type(focus) == sg.PySimpleGUI.Input:
            focus.update('б')

        if event == 'Left:37' and type(focus) == sg.PySimpleGUI.Input:
            focus.update('н')

        if event == 'Up:38':
            prev_focus.set_focus()

        if event == 'Down:40':
            next_focus.set_focus()

        if event == 'Отметить' or event == 'Отметить_Enter':
            absent = [values[i] for i in range(len(kids))]
            logger.info({k: v for k in kids for v in absent})
            if not all(absent):
                if sg.Window('Вы уверены?', [
                    [sg.Text('Вы отметили не всех! Сохранить?'), sg.Button('Да'),
                     sg.Button('Нет')]]).read(close=True)[0] == 'Нет':
                    continue
            mark_absent(path, values['date'], absent)
            break

    window.close()


def mark_absent(path, day, absent):
    work_book = load_workbook(path)
    ws = work_book['Посещаемость']
    column = int(day) + 4
    for row in range(len(absent)):
        ws.cell(row=row + 16, column=column).value = absent[row]

    work_book.save(path)

def get_absent(path, day, kids):
    absent = []
    work_book = load_workbook(path)
    ws = work_book['Посещаемость']
    column = int(day) + 4
    for i in range(len(kids)):
        if not ws.cell(row=i + 16, column=column).value:
            absent.append('')
            continue
        absent.append(ws.cell(row=i + 16, column=column).value)
    return absent

