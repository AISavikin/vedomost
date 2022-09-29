import PySimpleGUI as sg
from openpyxl import load_workbook
from utils import get_kids
from conf import *
from pathlib import Path


def kids_window(file_name: str):
    """
    Создает окно для добавления новых детей в ведомость.

    :param file_name: str: имя файла без расширения
    """
    # Путь к файлу
    path = Path(Path.cwd(), 'Ведомости', file_name)

    # Получаем имена и фамилии детей в виде списка строк

    kids = get_kids(path)

    # Левая колонка
    left_col = [
        [sg.Text('Фамилия и имя')],
        [sg.Input(key='name', focus=True, size=29)],
        [sg.Button('Добавить'), sg.Button('Удалить'), sg.Button('Исправить')],
        [sg.Button('Сохранить', expand_x=True)],
    ]
    # Правая колонка
    right_col = [
        [sg.Text(file_name)],
        [sg.Table(headings=['№', 'Имя'], values=list(enumerate(kids, 1)), key='table', col_widths=[7, 20],
                  auto_size_columns=False, justification='center', size=(20, 15))]
    ]

    layout = [[sg.Column(left_col, element_justification='c'), sg.Column(right_col, element_justification='c')]]
    window = sg.Window(f'Добавить ученика в группу {file_name}', layout, finalize=True,
                       font=(FONT_FAMILY, FONT_SIZE))
    window.bind("<Return>", "Добавить")

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break

        if event == 'Добавить' and values['name']:
            kids.append(values['name'])
            window['name'].update('')
            window['table'].update(values=enumerate(kids, 1))

        if event == 'Исправить' and values['table']:
            if not values['name']:
                sg.Popup('Введите имя!')
                continue
            kids[values['table'][0]] = values['name']
            window['table'].update(values=enumerate(kids, 1))
            window['name'].update('')

        if event == 'Удалить':
            kid = values['table'][0]
            kids.pop(kid)
            window['table'].update(values=enumerate(kids, 1))
            window['name'].update('')

        if event == 'Сохранить':
            save_new_kids(path, kids)
            break

    window.close()


def save_new_kids(path, kids):
    work_book = load_workbook(path)
    ws = work_book['Посещаемость']
    cells = 'B16:B38'
    kids += ['' for _ in range(23 - len(kids))]

    for i in range(len(kids)):
        ws[cells][i][0].value = kids[i]
    work_book.save(path)


if __name__ == '__main__':
    pass
