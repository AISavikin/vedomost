import PySimpleGUI as sg
import shutil
from utils import MONTH_DICT, date, get_work_days
from openpyxl import load_workbook, styles
from conf import *
from calendar import Calendar
from pathlib import Path


def add_new_sheet(file_name: str, list_group: list):
    """
    Создает окно для добавления новой ведомости
    :param file_name: str имя файла без расширения
    :param list_group: list список существующих ведомостей
    """
    # Добавляем в список групп значение "Новая группа"
    if 'Новая группа' not in list_group:
        list_group.append('Новая группа')

    left = [
        [sg.Text('На основе')],
        [sg.Text('Название ведомости')],
        [sg.Text('Месяц')]
    ]
    right = [
        [sg.Combo(list_group, size=(18, 0), key='base', default_value=file_name, readonly=True)],
        [sg.Input(size=(18, 0), key='file_name')],
        [sg.Combo(list(MONTH_DICT.keys()), size=(18, 0), key='month', default_value=f'{date:%B}', readonly=True)]
    ]
    layout = [
        [sg.Column(left), sg.Column(right)],
        [sg.Button('Добавить ведомость', key='add'), sg.Button('Отмена')]
    ]

    window = sg.Window('Добавить новую ведомость', layout, element_justification='c', font=(FONT_FAMILY, FONT_SIZE))

    while True:
        event, values = window.read()

        if event == 'Отмена' or event == sg.WINDOW_CLOSED:
            break

        if event == 'add':
            month = values['month']
            base = values['base']
            file_name = values['file_name']

            if not file_name and base == 'Новая группа':
                sg.Popup('Введите название группы!', title='Ошибка')
                continue
            status = create_new_sheet(file_name, base, month)

            if not status:
                sg.Popup('Такая группа уже есть', title='Ошибка')
                continue
            if sg.Window("Добавить ещё",
                         [[sg.Text(f"Ведомость {status.name} добавлена. Добавить ещё?")], [sg.Yes(), sg.No()]],
                         element_justification='c', modal=True).read(close=True)[0] == "Yes":
                continue
            else:
                break

    window.close()


# region NEW SHEET
def create_new_sheet(file_name, base, month):
    path = copy_sheet(file_name, base, month)
    if not path:
        return
    group = file_name.split('_')[0]
    clear_absent(path)
    clear_note(path)
    write_service_information(path, month, group)
    colorize_weekend(path, month)
    return path


def copy_sheet(file_name, base, month):
    if base == 'Новая группа':
        path_template = Path(Path.cwd(), 'Template.xlsx')
        new_path = Path(Path.cwd(), 'Ведомости', f'{file_name}_{month}.xlsx')
        shutil.copyfile(path_template, new_path)
        return Path(Path.cwd(), 'Ведомости', f'{file_name}_{month}.xlsx')
    file_name = base.split('_')[0]
    new_path = Path(Path.cwd(), 'Ведомости', f'{file_name}_{month}.xlsx')
    try:
        shutil.copyfile(f'Ведомости/{base}', new_path)
    except shutil.SameFileError:
        return
    return Path(Path.cwd(), 'Ведомости', f'{file_name}_{month}.xlsx')


def clear_note(path):
    wb = load_workbook(path)
    ws = wb['Заметки']
    cells = 'A1:A20'
    for i in ws[cells]:
        i[0].value = ''
    cells = 'B1:B20'
    for i in ws[cells]:
        i[0].value = ''

    wb.save(path)


def clear_absent(path):
    # Очищает ячейки от "н, б"
    # 16 - 39 диапазон строк, 5 - 36 диапазон столбцов
    work_book = load_workbook(path)
    ws = work_book['Посещаемость']
    for row in range(16, 39):
        for col in range(5, 36):
            ws.cell(row=row, column=col).value = ''
            ws.cell(row=row, column=col).fill = styles.PatternFill()

    work_book.save(path)


def colorize_weekend(path, month):
    work_book = load_workbook(path)
    ws = work_book['Посещаемость']
    weekends = [day[0] for day in Calendar().itermonthdays2(MONTH_DICT[month][1], MONTH_DICT[month][0]) if
                day[0] != 0 and day[1] in (5, 6)]
    for row in range(16, 39):
        for col in weekends:
            ws.cell(row=row, column=col + 4).fill = styles.PatternFill(start_color='5E5E5E', fill_type='solid')

    work_book.save(path)


def write_service_information(path, month, group):
    work_book = load_workbook(path)
    ws = work_book['Посещаемость']
    ws['N3'].value = month
    ws['AA42'].value = month
    ws['C5'].value = group
    ws['V3'].value = MONTH_DICT[month][1]
    ws['AG42'].value = MONTH_DICT[month][1]
    ws = work_book['Заметки']
    work_days = get_work_days(month)
    row = 1
    for day in work_days:
        ws.cell(row=row, column=1).value = day
        row += 1
    work_book.save(path)


# endregion

if __name__ == '__main__':
    pass
