import PySimpleGUI as sg
from openpyxl import load_workbook
from conf import *
from pathlib import Path

def notes_window(file_name: str):

    path = Path(Path.cwd(), 'Ведомости', file_name)

    notes = read_notes(path)


    l_notes = {day: notes[day] for num, day in enumerate(notes, 1) if num % 2 == 1}
    r_notes = {day: notes[day] for num, day in enumerate(notes, 1) if num % 2 == 0}

    l_col = [[sg.Text(day),
              sg.Multiline(default_text=l_notes[day], size=(50, 6), key=day,
                           no_scrollbar=True, font=(FONT_FAMILY, FONT_SIZE))] for day in l_notes]

    r_col = [[sg.Text(day),
              sg.Multiline(default_text=r_notes[day], size=(50, 6), key=day,
                           no_scrollbar=True, font=(FONT_FAMILY, FONT_SIZE))] for day in r_notes]

    layout = [
        [sg.Text(file_name, font=(FONT_FAMILY, FONT_SIZE + 5, 'underline')), sg.Button('Сохранить')],
        [sg.Column(l_col, scrollable=False, size_subsample_width=1, element_justification='r', size_subsample_height=1),
         sg.Column(r_col, scrollable=False, size_subsample_width=1, element_justification='r', size_subsample_height=1),
         ],
    ]

    window = sg.Window('Заметки', layout, element_justification='center', font=(FONT_FAMILY, FONT_SIZE), resizable=True)

    while True:
        event, values = window.read()

        if event == 'Отмена' or event == sg.WINDOW_CLOSED:
            break

        if event == 'Сохранить':
            notes = {day: values[day] for day in sorted(values)}
            save_notes(path, notes)
            break

    window.close()


def read_notes(path):
    workbook = load_workbook(path)
    ws = workbook['Заметки']
    row = 1
    notes = {}
    while ws.cell(row=row, column=1).value:
        if ws.cell(row=row, column=2).value:
            notes[ws.cell(row=row, column=1).value] = ws.cell(row=row, column=2).value
        else:
            notes[ws.cell(row=row, column=1).value] = ''

        row += 1
    return notes


def save_notes(path, notes):
    workbook = load_workbook(path)
    ws = workbook['Заметки']
    row = 1
    for day in notes:
        ws.cell(row=row, column=1).value = day
        ws.cell(row=row, column=2).value = notes[day]
        row += 1

    workbook.save(path)
