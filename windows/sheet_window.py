import PySimpleGUI as sg
from database import Student, Attendance
from utils import get_month_name, MONTH_DICT, date
from openpyxl import load_workbook, styles
from calendar import Calendar


def gen_headers(num_group, month):
    kids = Student.select().filter(group=num_group)
    kids_id = [str(i.id) for i in kids]
    data = Attendance.select().filter(month=month)
    days = set(i.day for i in data if str(i.id) in kids_id)
    if not days:
        return ['№', 'Фамилия, Имя', 'Н/д']
    return ['№', 'Фамилия, Имя'] + [str(i) for i in sorted(days)] + ['Н', 'Б']


def gen_table(num_group, month):
    kids = Student.select().filter(group=num_group)
    values = []
    for kid in kids:
        name = [kid.name]
        row = [(i.day, i.absent) for i in Attendance.select().filter(month=month, id=kid.id)]
        row = [i[1] for i in sorted(row)]
        values.append(name + row + [row.count('н'), row.count('б')])
    return values


def num_table(val):
    return [[i[0]] + i[1] for i in enumerate(val, 1)]


def sheet(num_group, month):
    month = MONTH_DICT[month][0]
    head = gen_headers(num_group, month)
    val = gen_table(num_group, month)
    groups = list(set(i.group for i in Student.select()))
    groups.remove(num_group)
    col_width = [3, 20] + [4 for _ in range(len(head))]

    layout = [
        [sg.Text(f'Группа {num_group} {get_month_name(month)}'), sg.Button('+'), sg.Combo(groups, k='group'),
         sg.Button('Закрыть ведомость')],
        [sg.Table(headings=head, values=num_table(val), col_widths=col_width, auto_size_columns=False,
                  justification='c', key='table')]
    ]

    window = sg.Window('Ведомость', layout, element_justification='c')

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        if event == '+':
            val = gen_table(num_group, month) + gen_table(values['group'], month)
            window['table'].Update(num_table(val))
        if event == 'Закрыть ведомость':
            close_day = sg.Window('', [[sg.Text('День закрытия ведомости')],
                                       [sg.Combo(list(range(1, 31)), default_value=date.day),
                                        sg.Text(get_month_name(month))],
                                       [sg.Button('Выбрать')]]).read(close=True)[1][0]
            close_sheet(num_group, values['group'], month, head, val, close_day)
    window.close()


def write_xls(head, val, file_name):
    wb = load_workbook('Template.xlsx')
    ws = wb['Посещаемость']
    cols = [2] + [int(i) + 4 for i in head[2:-2]] + [36, 38]

    for row in range(len(val)):
        data_row = val[row]
        for indx, column in enumerate(cols):
            ws.cell(row=row + 16, column=column).value = data_row[indx]
    wb.save(file_name)


def gen_file_name(num_group, ext_group, month):
    file_name = f'Ведомости/Группа {str(num_group)},{str(ext_group)} {get_month_name(month)}.xlsx'
    if not ext_group:
        file_name = file_name.replace(',', '')
    return file_name


def write_service_information(file_name, month, group, close_day):
    work_book = load_workbook(file_name)
    ws = work_book['Посещаемость']
    ws['N3'].value = month
    ws['AA42'].value = month
    ws['C5'].value = group
    ws['V3'].value = MONTH_DICT[month][1]
    ws['AG42'].value = MONTH_DICT[month][1]
    ws['Y42'].value = close_day
    work_book.save(file_name)

def colorize_weekend(file_name, month):
    work_book = load_workbook(file_name)
    ws = work_book['Посещаемость']
    name_month = get_month_name(month)
    weekends = [day[0] for day in Calendar().itermonthdays2(MONTH_DICT[name_month][1], MONTH_DICT[name_month][0]) if
                day[0] != 0 and day[1] in (5, 6)]
    for row in range(16, 39):
        for col in weekends:
            ws.cell(row=row, column=col + 4).fill = styles.PatternFill(start_color='5E5E5E', fill_type='solid')

    work_book.save(file_name)

def close_sheet(num_group, ext_group, month, head, val, close_day):
    file_name = gen_file_name(num_group, ext_group, month)
    write_xls(head, val, file_name)
    group = f'Группа {num_group}' if not ext_group else f'Группа {num_group}, {ext_group}'
    write_service_information(file_name, get_month_name(month), group, close_day)
    colorize_weekend(file_name, month)
    sg.Popup(f'Ведомость для {group} создана успешно!')



