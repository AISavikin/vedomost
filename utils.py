from openpyxl import load_workbook, styles
from datetime import datetime
from calendar import Calendar
from conf import *
from pathlib import Path
import shutil
import os

MONTH_DICT = {
    'Сентябрь': (9, YEAR[0]), 'Октябрь': (10, YEAR[0]), 'Ноябрь': (11, YEAR[0]), 'Декабрь': (12, YEAR[0]),
    'Январь': (1, YEAR[1]), 'Февраль': (2, YEAR[1]), 'Март': (3, YEAR[1]), 'Апрель': (4, YEAR[1]), 'Май': (5, YEAR[1])
}

date = datetime.now()


def get_work_days(month):
    wd = (2, 4)  # Рабочие дни (СРЕДА, ПЯТНИЦА)
    return [day[0] for day in Calendar().itermonthdays2(MONTH_DICT[month][1], MONTH_DICT[month][0])
            if day[0] != 0 and day[1] in wd]


def get_kids(path):
    """
    :param path: Относительный путь к файлу.
    :return: Список строк с именами детей
    """
    wb = load_workbook(path)
    ws = wb['Посещаемость']
    names = ws['B16:B38']
    return [names[i][0].value for i in range(23) if names[i][0].value]


# region NEW SHEET
def create_new_sheet(file_name, base, month):
    path = copy_sheet(file_name, base, month)
    if not path:
        return
    group = path.name.split('_')[0]
    clear_absent(path)
    clear_note(path)
    write_service_information(path, month, group)
    colorize_weekend(path, month)
    return path


def copy_sheet(file_name, base, month):
    if base == 'Новая группа':
        path_template = Path(Path.cwd(), 'Template.xlsx')
        new_path = Path(Path.cwd(), 'Ведомости', f'{file_name}_{month}.xlsx')
        shutil.copyfile(path_template, new_path)
        return Path(Path.cwd(), 'Ведомости', f'{file_name}_{month}.xlsx')
    file_name = base.split('_')[0]
    new_path = Path(Path.cwd(), 'Ведомости', f'{file_name}_{month}.xlsx')
    try:
        shutil.copyfile(f'Ведомости/{base}', new_path)
    except shutil.SameFileError:
        return
    return Path(Path.cwd(), 'Ведомости', f'{file_name}_{month}.xlsx')


def clear_note(path):
    wb = load_workbook(path)
    ws = wb['Заметки']
    cells = 'A1:A20'
    for i in ws[cells]:
        i[0].value = ''
    cells = 'B1:B20'
    for i in ws[cells]:
        i[0].value = ''

    wb.save(path)


def clear_absent(path):
    # Очищает ячейки от "н, б"
    # 16 - 39 диапазон строк, 5 - 39 диапазон столбцов
    work_book = load_workbook(path)
    ws = work_book['Посещаемость']
    ws['C7'].value = ''
    for row in range(16, 39):
        for col in range(5, 39):
            ws.cell(row=row, column=col).value = ''
            ws.cell(row=row, column=col).fill = styles.PatternFill()

    work_book.save(path)


def colorize_weekend(path, month):
    work_book = load_workbook(path)
    ws = work_book['Посещаемость']
    weekends = [day[0] for day in Calendar().itermonthdays2(MONTH_DICT[month][1], MONTH_DICT[month][0]) if
                day[0] != 0 and day[1] in (5, 6)]
    for row in range(16, 39):
        for col in weekends:
            ws.cell(row=row, column=col + 4).fill = styles.PatternFill(start_color='5E5E5E', fill_type='solid')

    work_book.save(path)


def write_service_information(path, month, group):
    work_book = load_workbook(path)
    ws = work_book['Посещаемость']
    ws['N3'].value = month
    ws['AA42'].value = month
    ws['C5'].value = group
    ws['V3'].value = MONTH_DICT[month][1]
    ws['AG42'].value = MONTH_DICT[month][1]
    ws = work_book['Заметки']
    work_days = get_work_days(month)
    row = 1
    for day in work_days:
        ws.cell(row=row, column=1).value = day
        row += 1
    work_book.save(path)


# endregion


def close_sheet(path):
    wb = load_workbook(path)
    ws = wb['Посещаемость']
    kids = get_kids(path)
    absents = [[cell.value for cell in ws["E16:AI38"][kid]] for kid in range(len(kids))]
    ws['C7'].value = len([day for day in absents[0] if day])
    for kid in range(len(kids)):
        ws['AJ16:AL38'][kid][0].value = absents[kid].count('н')
        ws['AJ16:AL38'][kid][2].value = absents[kid].count('б')
    wb.save(path)


def move_to_archive(path, month):
    if not os.path.exists('Ведомости/Архив'):
        os.mkdir('Ведомости/Архив')
    if not os.path.exists(f'Ведомости/Архив/{month}'):
        os.mkdir(f'Ведомости/Архив/{month}')
    shutil.move(path, Path('Ведомости', 'Архив', month, path.name))


def close_all_sheets():
    months = list(MONTH_DICT.keys())
    files = [file for file in Path('Ведомости').iterdir() if file.is_file() and file.name.endswith('.xlsx')]
    for file in files:
        month = file.name.split('_')[-1][:-5]
        close_sheet(file)
        if month != "Май":
            new_month = months[months.index(month) + 1]
            create_new_sheet('', file.name, new_month)
            move_to_archive(file, month)


if __name__ == '__main__':
    close_all_sheets()
