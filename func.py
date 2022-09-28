from calendar import Calendar
import os
from openpyxl import load_workbook, styles
import shutil
from datetime import datetime

date = datetime.now()


MONTH_DICT = {
    'Сентябрь': 9, 'Октябрь': 10, 'Ноябрь': 11, 'Декабрь': 12, 'Январь': 1, 'Февраль': 2, 'Март': 3, 'Апрель': 4,
    'Май': 5
}

def get_kids(file_name):
    """
    :param file_name: Относительный путь к файлу
    :return: Список строк с именами детей
    """
    wb = load_workbook(f'Ведомости/{file_name}')
    ws = wb['Посещаемость']
    names = ws['B16:B38']
    return [names[i][0].value for i in range(23) if names[i][0].value]


def save_new_kids(file_name, kids):
    file_name = f'Ведомости/{file_name}'
    work_book = load_workbook(file_name)
    ws = work_book['Посещаемость']
    cells = 'B16:B38'

    for i in ws[cells]:
        i[0].value = ''

    for i in range(len(kids)):
        ws[cells][i][0].value = kids[i]

    work_book.save(file_name)


def clear_absent(file_name):
    file_name = f'Ведомости/{file_name}'
    # Очищает ячейки от "н, б"
    # 16 - 39 диапазон строк, 5 - 36 диапазон столбцов
    work_book = load_workbook(file_name)
    ws = work_book['Посещаемость']
    for row in range(16, 39):
        for col in range(5, 36):
            ws.cell(row=row, column=col).value = ''
            ws.cell(row=row, column=col).fill = styles.PatternFill()

    work_book.save(file_name)


def mark_absent(file_name, day, absent):
    file_name = f'Ведомости/{file_name}'
    work_book = load_workbook(file_name)
    ws = work_book['Посещаемость']
    column = int(day) + 4
    for row in range(len(absent)):
        ws.cell(row=row + 16, column=column).value = absent[row]

    work_book.save(file_name)


def read_notes(file_name):
    workbook = load_workbook(f'Ведомости/{file_name}')
    ws = workbook['Заметки']
    row = 1
    notes = {}
    while ws.cell(row=row, column=1).value:
        if ws.cell(row=row, column=2).value:
            notes[ws.cell(row=row, column=1).value] = ws.cell(row=row, column=2).value
        else:
            notes[ws.cell(row=row, column=1).value] = ''

        row += 1
    return notes


def save_notes(file_name, notes):
    file_name = f'Ведомости/{file_name}'
    workbook = load_workbook(file_name)
    ws = workbook['Заметки']
    row = 1
    for day in notes:
        ws.cell(row=row, column=1).value = day
        ws.cell(row=row, column=2).value = notes[day]
        row += 1

    workbook.save(file_name)


# region NEW SHEET
def create_new_sheet(file_name, base, month):
    file_name = copy_sheet(file_name, base, month)
    if not file_name:
        return
    group = file_name.split('_')[0]
    clear_absent(file_name)
    clear_note(file_name)
    write_service_information(file_name, month, group)
    colorize_weekend(file_name, month)
    return 'OK'


def copy_sheet(file_name, base, month):
    if base == 'Новая группа':
        shutil.copyfile('Template.xlsx', f'Ведомости/{file_name}_{month}.xlsx')
        return f'{file_name}_{month}.xlsx'
    file_name = base.split('_')[0]
    new_path = f'Ведомости/{file_name}_{month}.xlsx'
    try:
        shutil.copyfile(f'Ведомости/{base}', new_path)
    except shutil.SameFileError:
        return
    return f'{file_name}_{month}.xlsx'


def clear_note(file_name):
    file_name = f'Ведомости/{file_name}'
    wb = load_workbook(file_name)
    ws = wb['Заметки']
    cells = 'A1:A20'
    for i in ws[cells]:
        i[0].value = ''
    cells = 'B1:B20'
    for i in ws[cells]:
        i[0].value = ''

    wb.save(file_name)


def colorize_weekend(file_name, month):
    file_name = f'Ведомости/{file_name}'
    work_book = load_workbook(file_name)
    ws = work_book['Посещаемость']
    weekends = [day[0] for day in Calendar().itermonthdays2(2022, MONTH_DICT[month]) if
                day[0] != 0 and day[1] in (5, 6)]
    for row in range(16, 39):
        for col in weekends:
            ws.cell(row=row, column=col + 4).fill = styles.PatternFill(start_color='5E5E5E', fill_type='solid')

    work_book.save(file_name)


def get_work_days(month, year=2022, wd=(2, 4)):
    c = Calendar()
    return [day[0] for day in c.itermonthdays2(year, month) if day[0] != 0 and day[1] in wd]


def write_service_information(file_name, month, group):
    file_name = f'Ведомости/{file_name}'
    work_book = load_workbook(file_name)
    ws = work_book['Посещаемость']
    ws['N3'].value = month
    ws['AA42'].value = month
    ws['C5'].value = group
    ws = work_book['Заметки']
    work_days = get_work_days(MONTH_DICT[month])
    row = 1
    for day in work_days:
        ws.cell(row=row, column=1).value = day
        row += 1
    work_book.save(file_name)
# endregion

if __name__ == '__main__':
    print(get_kids('Группа 1_Октябрь.xlsx'))
